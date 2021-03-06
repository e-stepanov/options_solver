# -*- coding: utf-8 -*-
""" Explicit finite difference scheme for asian options """

import time

import numpy as np
import matplotlib.pyplot as plt
from matplotlib import cm
from openpyxl import (
    Workbook,
    load_workbook
)
from mpl_toolkits.mplot3d import Axes3D

from ..core import FDMBase


class AsianOptionExplicitFDM(FDMBase):
    """
    Explicit Euler scheme realization for Black-Scholes PDE
    for vanilla europian option
    """

    def __init__(self, option, market, nodes):
        # fdm parameters
        super(AsianOptionExplicitFDM, self).__init__(option, market, nodes)

        self._t_nodes = self.nodes.time_nodes
        self._S_nodes = self.nodes.asset_price_nodes
        self._A_nodes = self.nodes.average_price_nodes

        self._t_number = len(self.nodes.time_nodes)
        self._S_number = len(
            self.nodes.asset_price_nodes
        )
        self._A_number = len(
            self.nodes.average_price_nodes
        )

        self.dt = (
            self.nodes.time_nodes[1] - self.nodes.time_nodes[0]
        )
        self.dS = (
            self.nodes.asset_price_nodes[1] -
            self.nodes.asset_price_nodes[0]
        )
        self.dA = (
            self.nodes.average_price_nodes[1] -
            self.nodes.average_price_nodes[0]
        )

    # BOUNDARY VALUES

    def get_boundary_left(self, time_node):
        """ Boundary values for S = 0 """
        return (
            np.exp(-self.market.interest * time_node) *
            self.option.calculate_payoff(self._A_nodes)
        )

    def get_boundary_right(self, time_node):
        """ Boundary values for S = S_max """
        return (
            np.maximum(
                np.exp(-self.market.interest * time_node) *
                (self._A_nodes / self.option.maturity -
                 self.option.strike) +
                self._S_nodes[-1] /
                (self.market.interest * self.option.maturity) *
                (1.0 - np.exp(-self.market.interest * time_node)), 0
            )
        )

    def get_boundary_front(self, C_j1):
        """ Boundary values for A = 0 """
        return C_j1

    def get_boundary_back(self, time_node):
        """ Boundary values for A = A_max"""
        return (
            np.exp(-self.market.interest * time_node) *
            self.option.calculate_payoff(
                self._A_nodes[-1]
            ) +
            self._S_nodes / (self.market.interest * self.option.maturity) *
            (1 - np.exp(-self.market.interest * time_node))
        )

    # INITIAL VALUES
    def get_initial(self):
        """ Initial values of option prices (tau = 0) """
        return np.tile(
            self.option.calculate_payoff(self._A_nodes),
            (self._S_number, 1)
        )

    # COEFFICIENTS OF FINITE DIFFERENCE SCHEME
    def get_coeffs_center(self):
        """ Coefficients for C_jk """
        return (
            1 - self.dt * (np.arange(
                self._S_number)**2 *
                self.market.volatility**2 / 2.0 + self.market.interest)
        )

    def get_coeffs_right(self):
        """ Coefficients for C_{j+1,k} """
        S_range = np.arange(self._S_number)
        return (
            self.dt / 2.0 * (S_range**2 * self.market.volatility**2 / 2.0 +
                             S_range * self.market.interest)
        )

    def get_coeffs_left(self):
        """ Coefficients for C_{j-1,k} """
        S_range = np.arange(self._S_number)
        return self.dt / 2.0 * (S_range**2 * self.market.volatility**2 / 2.0 -
                                S_range * self.market.interest)

    def get_coeffs_front(self):
        """ Coefficients for C_{j,k+1} """
        return (
            np.arange(self._S_number) *
            self.dS * self.dt / (2.0 * self.dA)
        )

    def get_coeffs_back(self):
        """ Coefficients for C_{j,k-1} """
        return (
            -np.arange(self._S_number) *
            self.dS * self.dt / (2.0 * self.dA)
        )

    def export_to_file(self, filename):
        """ Write computed values to file """
        zero_volatility_solution = self.get_zero_volatility_solution()

        export_data = (
            [
                self._S_nodes[i],
                self._A_nodes[j],
                self.option_prices[i, j],
                zero_volatility_solution[i, j]
            ] for i in range(self._S_number) for j in range(self._A_number)
        )

        wb = Workbook()
        ws = wb.active

        ws['A1'] = 'Asset price'
        ws['B1'] = 'Average price'
        ws['C1'] = 'Numerical price'
        ws['D1'] = 'Zero volatility price'
        ws['E1'] = 'Difference'

        row = 2
        for (
            asset_price, average_price,
            numerical_price, zero_volatility_price
        ) in export_data:
            ws.cell(row=row, column=1).value = asset_price
            ws.cell(row=row, column=2).value = average_price
            ws.cell(row=row, column=3).value = numerical_price
            ws.cell(row=row, column=4).value = zero_volatility_price
            ws.cell(row=row, column=5).value = (
                numerical_price - zero_volatility_price
            )
            row += 1

        wb.save(filename)

    @staticmethod
    def plot_difference_from_file(filename):
        wb = load_workbook(filename)
        ws = wb.active
        asset_prices = [0] * (ws.max_row - 1)
        average_prices = [0] * (ws.max_row - 1)
        differences = [0] * (ws.max_row - 1)

        index = 0
        for row in ws.iter_rows('A2:E%d' % ws.max_row):
            asset_prices[index] = row[0].value
            average_prices[index] = row[1].value
            differences[index] = row[4].value
            index += 1

        asset_prices = np.array(
            sorted(list(set(asset_prices)))
        )
        average_prices = np.array(
            sorted(list(set(average_prices)))
        )

        average_prices_grid, asset_prices_grid = np.meshgrid(
            average_prices, asset_prices
        )
        print len(differences)
        differences_grid = np.array(differences).reshape(
            asset_prices_grid.shape
        )

        figure = plt.figure()
        axes = Axes3D(figure)
        axes.plot_surface(
            asset_prices_grid,
            average_prices_grid,
            differences_grid,
            rstride=1, cstride=1, cmap=cm.YlGnBu_r
        )
        plt.show()

    def get_zero_volatility_solution(self):
        """ Return precise solution for zero volatility at tau = T """
        A_grid, S_grid = np.meshgrid(self._A_nodes, self._S_nodes)
        return np.maximum(
            (A_grid / self.option.maturity - self.option.strike) *
            np.exp(-self.market.interest * self.option.maturity) +
            S_grid / (self.market.interest * self.option.maturity) *
            (1 - np.exp(-self.market.interest * self.option.maturity)), 0
        )

    def calculate_prices(self):
        start = time.time()
        print("Begin: %s" % str(time.time() - start))

        coeffs_center = self.get_coeffs_center()
        coeffs_right = self.get_coeffs_right()
        coeffs_left = self.get_coeffs_left()
        coeffs_front = self.get_coeffs_front()
        coeffs_back = self.get_coeffs_back()

        print(
            "Coeffs were counted succesfully: %s" %
            str(time.time() - start)
        )

        C_next = np.zeros((self._S_number, self._A_number))
        C_current = self.get_initial()

        for time_node in self.nodes.time_nodes[1:]:
            # every 100th iteration print time info
            if not int(time_node * self._t_number) % 100:
                print(
                    "tau = %s. Time elapsed: %s" %
                    (time_node, str(time.time() - start))
                )
                print(
                    "Minimum value %f" % np.min(C_current)
                )

            # values inside the area
            C_next[1:self._S_number - 1, 1:self._A_number - 1] = (
                C_current[1:self._S_number - 1,
                          1:self._A_number - 1] *
                coeffs_center[1:self._S_number - 1, np.newaxis] +
                C_current[0:self._S_number - 2,
                          1:self._A_number - 1] *
                coeffs_left[1:self._S_number - 1, np.newaxis] +
                C_current[2:self._S_number, 1:self._A_number - 1] *
                coeffs_right[1:self._S_number - 1, np.newaxis] +
                C_current[1:self._S_number - 1,
                          0:self._A_number - 2] *
                coeffs_back[1:self._S_number - 1, np.newaxis] +
                C_current[1:self._S_number - 1, 2:self._A_number] *
                coeffs_front[1:self._S_number - 1, np.newaxis]
            )

            # boundary values
            C_next[0] = (  # S = 0
                self.get_boundary_left(time_node)
            )
            C_next[self._S_number - 1] = (  # S = S_max
                self.get_boundary_right(time_node)
            )
            C_next[:, 0] = (  # A = 0
                self.get_boundary_front(C_next[:, 1])
            )
            C_next[:, self._A_number - 1] = (  # A = A_max
                self.get_boundary_back(time_node)
            )
            C_current = C_next.copy()

        self.option_prices = C_current

        return C_current

    def plot_option_prices(self, asset_price_sparse=1, average_price_sparse=1):
        """
        Plot option prices with respect to initial asset price and
        average asset price
        """
        average_prices_grid, asset_prices_grid = np.meshgrid(
            self.nodes.average_price_nodes[::average_price_sparse],
            self.nodes.asset_price_nodes[::asset_price_sparse]
        )
        figure = plt.figure()
        axes = Axes3D(figure)
        axes.plot_surface(
            asset_prices_grid, average_prices_grid,
            self.option_prices[::asset_price_sparse,
                               ::average_price_sparse],
            rstride=1, cstride=1, cmap=cm.YlGnBu_r
        )
        plt.show()
